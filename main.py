from selenium import webdriver
from selenium.webdriver.firefox.options import Options  # for headless
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
from bs4 import BeautifulSoup
import requests
import urllib3
import time

links_list = []


def main():
    options = Options()
    #options.headless = True
    browser = webdriver.Firefox(options=options)
    browser.get('https://www.inmyroom.ru/products/myagkaya-mebel/divany/pryamye-divany')
    browser.find_element_by_class_name('b-CategoriesPreviews_showMore').click()  # открыть полный список нужных товаров


    wb = openpyxl.load_workbook(filename='sample.xlsx')
    ws = wb.active

    ws['A1'] = 'Название'
    ws['B1'] = 'Цена'
    ws['C1'] = 'Категория'
    ws['D1'] = 'Ширина'
    ws['E1'] = 'Высота'
    ws['F1'] = 'Глубина'
    ws['G1'] = 'Бренд'
    ws['H1'] = 'Страна'
    ws['I1'] = 'Артикул'
    elem = browser.find_element_by_class_name('pb-CategoryTeasers')
    for element in elem.find_elements_by_class_name('b-CategoryPreview'):
        link_to_category = str(element.find_element_by_class_name('b-CategoryPreview_title').get_property('href'))
        #browser.execute_script("window.open('" + link_to_category + "')")  # открытие нового окна со списком товаров
        #browser.switch_to.window(browser.window_handles[1])
        time.sleep(5)
        total_pages = browser.find_elements_by_class_name('b-Pagination_item')
        if len(total_pages) > 1:
            pages_count: str = total_pages[-1].find_element_by_class_name('b-Pagination_link').get_property('href')[-2:]
            #  если количество всех страниц двузначное
            if not pages_count.startswith("/"):
                pages = pages_count
            else:
                pages = pages_count[1]
            for page in range(1, int(pages) + 1):
                window = link_to_category + '/page/' + str(page)

                browser.execute_script("window.open('" + window + "')")
                browser.switch_to.window(browser.window_handles[1])

                time.sleep(4)
                items = browser.find_elements_by_class_name('s-ProductPreview_b-TitleBlock_title')
                counter = 2
                for item in items:
                    browser.execute_script("window.open('" + item.get_property('href') + "')")
                    browser.switch_to.window(browser.window_handles[2])

                    time.sleep(6)
                    product_info = browser.find_element_by_class_name('s-ProductCard_b-Metrics_list')
                    ws['A' + str(counter)] = product_info.find_elements_by_class_name('s-ProductCard_b-Metrics_item')[0].find_element_by_class_name('s-ProductCard_b-Metrics_text').text
                    ws['B' + str(counter)] = browser.find_element_by_class_name('s-ProductCard_b-CurrentPrice')\
                                                   .find_element_by_tag_name('span').get_property('content')
                    ws['C' + str(counter)] = browser.find_elements_by_class_name('b-Breadcrumbs_item')[-1]\
                                                   .find_element_by_tag_name('span')\
                                                   .text
                    ws['D' + str(counter)] = product_info.find_elements_by_class_name('s-ProductCard_b-Metrics_item')[1].find_element_by_class_name('s-ProductCard_b-Metrics_text').text
                    ws['E' + str(counter)] = product_info.find_elements_by_class_name('s-ProductCard_b-Metrics_item')[2].find_element_by_class_name('s-ProductCard_b-Metrics_text').text
                    ws['F' + str(counter)] = product_info.find_elements_by_class_name('s-ProductCard_b-Metrics_item')[3].find_element_by_class_name('s-ProductCard_b-Metrics_text').text
                    ws['G' + str(counter)] = product_info.find_elements_by_class_name('s-ProductCard_b-Metrics_item')[5].find_element_by_class_name('s-ProductCard_b-Metrics_text').text
                    ws['H' + str(counter)] = product_info.find_elements_by_class_name('s-ProductCard_b-Metrics_item')[6].find_element_by_class_name('s-ProductCard_b-Metrics_text').text

                    ws['I' + str(counter)] = browser.find_element_by_class_name('s-ProductCard_b-Info_item')\
                                                    .find_element_by_tag_name('span').text

                    counter += 1

                    wb.save(filename="sample.xlsx")
                    browser.close()
                    browser.switch_to.window(browser.window_handles[1])
                browser.close()
                browser.switch_to.window(browser.window_handles[0])
        browser.close()
        browser.switch_to.window(browser.window_handles[0])


    wb.save(filename="sample.xlsx")
    browser.quit()


if __name__ == '__main__':
    main()